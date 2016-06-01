# Generates a field map when passed an appropriately formatted CSV with pre-filled row and range information
# www.github.com/slin63
# slin63@illinois.edu
from openpyxl import Workbook
from openpyxl.utils import _get_column_letter
from openpyxl.styles import PatternFill
from datetime import datetime


class PlotCell(object):
    def __init__(self, range_num, row_num, experiment, plot_id, field):
        self.range = range_num
        self.row = row_num
        self.experiment = experiment
        self.plot_id = plot_id
        self.field = field

    def __repr__(self):
        return self.plot_id


def compile_info(plot_objects):
    """
    :plot_objects: List containing selected ObsPlot objects
    :return: .xlsx containing a colored field map of the passed plots.
    """
    wb = Workbook()
    field_set = get_field_object_set(plot_objects)
    for field in field_set:
        field_plots = get_plots_in_field(plot_objects, field)
        worksheet = wb.create_sheet()
        worksheet.title = field.field_name
        experiment_current = field_plots[0].experiment
        experiment_colors = iter(['00ccff', '00ffcc', 'ffcccc'])
        cell_color = experiment_colors.next()

        for plot in field_plots:
            coordinate = plot.range + str(plot.row)
            if plot.experiment != experiment_current:
                experiment_current = plot.experiment
                # Cycles through three experiment colors so the spreadsheet doesn't look incredibly dull
                try:
                    cell_color = experiment_colors.next()
                except StopIteration:
                    experiment_colors = iter(['00ccff', '00ffcc', 'ffcccc'])
                    cell_color = experiment_colors.next()

            cell_fill = PatternFill(start_color=cell_color,
                       end_color=cell_color,
                       fill_type='solid')

            worksheet[coordinate] = plot.plot_id
            worksheet[coordinate].fill = cell_fill

        domain = get_plot_domains(field_plots)

        add_axes(worksheet, domain, field_plots)

    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))

    return wb


def add_axes(worksheet, domain, plot_objects):
    """
    :param worksheet: Worksheet we will be appending with information.
    :param domain: Rows and ranges.
    :return: Excel file with row and range axes.
    """
    axes = generate_axes(domain, plot_objects)
    for axis in axes:
        for coordinate in axis.keys():
            worksheet[coordinate] = axis[coordinate]

    worksheet['A1'] = 'FieldMapper / slin63@illinois.edu / FCPathology / Rendered: {}. Experiments described at bottom of sheet.'.format(datetime.now())

    return 0


def generate_axes(domain, plot_objects):
    """
    :param domain: Rows and ranges.
    :return: Dictionaries formatted {ExcelIndex (e.g. H23): Row or range value} to use as axes.
    """
    row_max = max(domain[0])
    row_min = min(domain[0])

    ranges = [letter_to_number(e) for e in domain[1]]
    range_max = _get_column_letter(max(ranges))
    range_min = _get_column_letter(min(ranges))

    range_min_sub_one = _get_column_letter(letter_to_number(range_min) - 1)
    range_max_plus_one = _get_column_letter(letter_to_number(range_max) + 1)
    row_min_sub_one = row_min - 1
    row_max_plus_one = row_max + 1

    labels = {range_min_sub_one + str(row_min_sub_one): 'Rows/Ranges'}
    experiments = get_plot_experiments(plot_objects)

    row_axes = {}
    for e in xrange(row_min, row_max + 1):
        row_axes[range_min_sub_one + str(e)] = e
        row_axes[range_max_plus_one + str(e)] = e

    range_axes = {}
    for e in xrange(letter_to_number(range_min), letter_to_number(range_max) + 1):
        range_axes[_get_column_letter(e) + str(row_min_sub_one)] = e
        range_axes[_get_column_letter(e) + str(row_max_plus_one)] = e

    experiment_axes = {}
    current_row = row_max + 2
    for exp in experiments:
        exp_string = 'EXP: {} - {}. Owner: {}. Field: {}. Purpose: {}. Comments: {}.'.format(exp.name, exp.start_date, exp.user, exp.field, exp.purpose, exp.comments)
        experiment_axes[range_min + str(current_row)] = exp_string
        current_row += 1


    return row_axes, range_axes, labels, experiment_axes


def empty_field():
    wb = Workbook()
    ws = wb.active
    ws['C3'] = 'No data for this field.'

    return wb

def letter_to_number(letter):
    letter = letter.lower()
    l_to_n = {
        'a': 1, 'b': 2, 'c': 3, 'd': 4, 'e': 5, 'f': 6, 'g': 7, 'h': 8, 'i': 9, 'j': 10, 'k': 11, 'l': 12, 'm': 13,
        'n': 14, 'o': 15, 'p': 16, 'q': 17, 'r': 18, 's': 19, 't': 20, 'u': 21, 'v': 22, 'w': 23, 'x': 24, 'y': 25,
        'z': 26, 'aa': 27, 'ab': 28, 'ac': 29, 'ad': 30, 'ae': 31, 'af': 32, 'ag': 33, 'ah': 34, 'ai': 35, 'aj': 36,
        'ak': 37, 'al': 38, 'am': 39, 'an': 40
    }
    return l_to_n[letter]


def get_field_object_set(object_list):
    field_set = set()
    for obj in object_list:
        field_set.add(obj.field)

    return field_set


def get_plots_in_field(object_list, field):
    plots_in_field = []
    for obj in object_list:
        if obj.field == field:
            plots_in_field.append(obj)

    return plots_in_field


def get_plot_domains(object_list):
    rows = []
    ranges = []
    for obj in object_list:
        rows.append(int(obj.row))
        ranges.append(obj.range)

    return [rows, ranges]


def get_plot_experiments(object_list):
    experiment_set = set()
    for obj in object_list:
        experiment_set.add(obj.experiment)

    return experiment_set

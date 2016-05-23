# https://openpyxl.readthedocs.org/en/default/
from openpyxl import load_workbook, Workbook


def get_index(start, end):
    """
    :param start: Excel index of the first messed up plot
    :param end: Excel index of the very last messed up plot
    :return: Set of indices to pass to swap_rows
    """
    indices = []
    index = -1
    count = 2
    length = end-start
    # Length in xrange loop is arbitrary
    for e in xrange(length):
        if index < length:
            index += 1
            if count % 4 == 0:
                index += 4
                indices.append(index)
                count += 1
            else:
                indices.append(index)
                count += 1

    true_indices = []

    for index in indices:
        index = 'H' + str(index + start)
        true_indices.append(index)

    # http://stackoverflow.com/questions/4647050/collect-every-pair-of-elements-from-a-list-into-tuples-in-python
    # Pairs off elements of the list in twos
    zipped = zip(true_indices[0::2], true_indices[1::2])

    return zipped


def swap_rows(indices, output):
    """
    :param indices: A tuple containing lists containing tuple pairs of indices whose Row values will be swapped
    :return: Excel file named `output.xlsx` containing updated rows.
    """
    wb = load_workbook('dbo_row_sheanlin_revision.xlsx')
    ws = wb['dbo_row_sheanlin_revision']

    for list in indices:
        print list

        for pair in list:
            # Swapping the values
            print 'Swapping: [', pair[0], ':',ws[pair[0]].value, '] and [', pair[1], ':', ws[pair[1]].value, ']'
            row_a = ws[pair[0]].value
            row_b = ws[pair[1]].value

            ws[pair[0]] = row_b
            ws[pair[1]] = row_a

            # print ws[pair[0]].value, ws[pair[1]].value

    wb.save(output)

    return 0

indices_a = get_index(2145, 2194)
indices_b = get_index(2259, 2308)
indices_c = get_index(2373, 2422)
indices_d = get_index(2749, 2798)
indices_e = get_index(2443, 2492)
indices_f = get_index(2557, 2606)

indices_g = get_index(2671, 2704)

swap_rows(
    (indices_a, indices_b, indices_c, indices_d, indices_e, indices_f, indices_g),
    "output.xlsx"
)





